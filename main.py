import argparse
import mimetypes
import os
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple
from urllib.parse import quote

from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

WHATSAPP_WEB_URL = "https://web.whatsapp.com"
DEFAULT_SHEET_NAME = "Data"
STATUS_COL = 5  # E
TIMESTAMP_COL = 6  # F


class WhatsAppAutomationError(Exception):
    """Raised for invalid configuration or runtime WhatsApp automation conditions."""


def now_str() -> str:
    return datetime.now().isoformat(sep=" ", timespec="seconds")


def sanitize_phone_number(raw: object, default_country_code: Optional[str] = None) -> Optional[str]:
    if raw is None:
        return None

    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return None

    if default_country_code:
        cc = re.sub(r"\D", "", str(default_country_code))
        if cc and len(digits) <= 10 and not digits.startswith(cc):
            digits = f"{cc}{digits}"

    return digits if len(digits) >= 10 else None


def wait_for_login(driver: webdriver.Chrome, timeout: int = 240) -> None:
    print("Open WhatsApp Web and scan QR if required...")
    WebDriverWait(driver, timeout).until(
        EC.any_of(
            EC.presence_of_element_located((By.XPATH, "//div[@aria-label='Chat list']")),
            EC.presence_of_element_located((By.XPATH, "//div[@role='grid']")),
            EC.presence_of_element_located((By.XPATH, "//div[@id='pane-side']")),
        )
    )
    print("WhatsApp login detected.")


def dismiss_invalid_number_popup(driver: webdriver.Chrome, timeout: int = 3) -> None:
    """Click OK on invalid-number popup if shown so script can proceed cleanly."""
    candidates = [
        "//button[normalize-space()='OK']",
        "//div[@role='button' and normalize-space()='OK']",
        "//button[.//span[normalize-space()='OK']]",
    ]
    for xp in candidates:
        try:
            btn = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xp)))
            btn.click()
            return
        except TimeoutException:
            continue


def open_chat(driver: webdriver.Chrome, phone: str, message: str, timeout: int = 60) -> Tuple[bool, str]:
    encoded = quote(message or "")
    driver.get(f"{WHATSAPP_WEB_URL}/send?phone={phone}&text={encoded}")

    invalid_xpaths = [
        "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'not on whatsapp')]",
        "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'phone number shared via url is invalid')]",
        "//*[contains(translate(text(),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'number is invalid')]",
    ]

    try:
        WebDriverWait(driver, timeout).until(
            EC.any_of(
                EC.presence_of_element_located((By.XPATH, "//footer//div[@contenteditable='true']")),
                *[EC.presence_of_element_located((By.XPATH, xp)) for xp in invalid_xpaths],
            )
        )
    except TimeoutException:
        dismiss_invalid_number_popup(driver)
        return False, "Timeout opening chat (number may be invalid/non-WhatsApp or network is slow)"

    for xp in invalid_xpaths:
        if driver.find_elements(By.XPATH, xp):
            dismiss_invalid_number_popup(driver)
            return False, "Number not on WhatsApp or invalid"

    return True, "Chat opened"


def send_text(driver: webdriver.Chrome, timeout: int = 20) -> Tuple[bool, str]:
    try:
        input_box = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable((By.XPATH, "//footer//div[@contenteditable='true']"))
        )
        input_box.click()
        input_box.send_keys("\n")
        return True, "Text sent"
    except TimeoutException:
        return False, "Text send failed: composer not ready"


def _set_attachment_caption(driver: webdriver.Chrome, caption: str, timeout: int) -> None:
    if not caption:
        return

    # In latest WhatsApp Web, caption may appear either in a dialog or in bottom composer.
    caption_xpaths = [
        "//div[@role='dialog']//div[contains(@aria-label,'caption') and @contenteditable='true']",
        "//div[@role='dialog']//div[@contenteditable='true' and @data-tab='10']",
        "//footer//div[@contenteditable='true' and (@data-tab='10' or @data-tab='1')]",
        "(//div[@contenteditable='true'])[last()]",
    ]

    for xp in caption_xpaths:
        try:
            caption_box = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((By.XPATH, xp)))
            caption_box.click()
            caption_box.send_keys(caption)
            return
        except Exception:
            continue

    raise TimeoutException("caption input not found")


def send_attachment(driver: webdriver.Chrome, file_path: str, caption: str = "", timeout: int = 45) -> Tuple[bool, str]:
    """Attach and send local file with optional caption in one message."""
    if not file_path:
        return True, "No attachment"

    resolved = Path(file_path).expanduser().resolve()
    if not resolved.exists() or not resolved.is_file():
        return False, f"Attachment not found: {resolved}"

    mime_type, _ = mimetypes.guess_type(str(resolved))
    media_file = bool(mime_type and (mime_type.startswith("image/") or mime_type.startswith("video/")))

    try:
        attach_button = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable(
                (
                    By.XPATH,
                    "//button[@title='Attach' or @aria-label='Attach' or .//span[@data-icon='plus'] or .//*[contains(@data-testid,'attach')] or .//*[contains(@data-icon,'plus')]]",
                )
            )
        )
        attach_button.click()

        # Try clicking menu entry first, but continue even if UI text changes.
        option_xpath = (
            "//*[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'photos') or contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'video')]/ancestor::*[@role='button' or self::li][1]"
            if media_file
            else "//*[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'document')]/ancestor::*[@role='button' or self::li][1]"
        )
        try:
            WebDriverWait(driver, 4).until(EC.element_to_be_clickable((By.XPATH, option_xpath))).click()
        except TimeoutException:
            pass

        file_inputs = WebDriverWait(driver, timeout).until(
            EC.presence_of_all_elements_located((By.XPATH, "//input[@type='file']"))
        )

        # First target file input tied to selected menu item; then fall back to all file inputs.
        preferred = []
        targeted_input_xpath = (
            "//*[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'photos') or contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'video')]//input[@type='file']"
            if media_file
            else "//*[contains(translate(normalize-space(.),'ABCDEFGHIJKLMNOPQRSTUVWXYZ','abcdefghijklmnopqrstuvwxyz'),'document')]//input[@type='file']"
        )
        preferred.extend(driver.find_elements(By.XPATH, targeted_input_xpath))

        fallback = []
        for el in file_inputs:
            if el not in preferred:
                fallback.append(el)

        candidates = preferred + fallback
        uploaded = False
        last_error = ""
        for file_input in candidates:
            try:
                driver.execute_script(
                    "arguments[0].style.display='block'; arguments[0].style.visibility='visible'; arguments[0].style.opacity=1;",
                    file_input,
                )
                file_input.send_keys(str(resolved))
                # Uploaded when preview appears (blob image/video or media stage controls).
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located(
                        (
                            By.XPATH,
                            "//img[contains(@src,'blob:')] | //video | //button[@aria-label='Send'] | //span[@data-icon='send']",
                        )
                    )
                )
                uploaded = True
                break
            except Exception as exc:
                last_error = str(exc)

        if not uploaded:
            return False, f"Attachment failed: could not upload via available file inputs ({last_error or 'unknown'})"

        # Caption may be in modal or in footer inline composer.
        # If caption box isn't found, continue (message may already be prefilled in chat before attach).
        try:
            _set_attachment_caption(driver, caption, 12)
        except TimeoutException:
            if caption:
                try:
                    footer_box = WebDriverWait(driver, 4).until(
                        EC.element_to_be_clickable((By.XPATH, "//footer//div[@contenteditable='true']"))
                    )
                    footer_box.click()
                    footer_box.send_keys(caption)
                except TimeoutException:
                    pass

        send_button = WebDriverWait(driver, timeout).until(
            EC.element_to_be_clickable(
                (By.XPATH, "//button[@aria-label='Send'] | //span[@data-icon='send']/ancestor::button")
            )
        )
        send_button.click()
        return True, "Attachment sent with caption"
    except TimeoutException:
        return False, "Attachment failed: timeout waiting for attachment UI"


def write_result(sheet, row: int, status: str) -> None:
    sheet.cell(row=row, column=STATUS_COL).value = status
    sheet.cell(row=row, column=TIMESTAMP_COL).value = now_str()


def process_rows(
    excel_path: Path,
    sheet_name: str,
    default_country_code: Optional[str],
    pause_seconds: float,
    cooldown_every: int,
    cooldown_seconds: float,
    open_excel_after: bool,
) -> None:
    keep_vba = excel_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(excel_path, keep_vba=keep_vba)
    if sheet_name not in workbook.sheetnames:
        raise WhatsAppAutomationError(f"Sheet '{sheet_name}' not found in {excel_path}")

    sheet = workbook[sheet_name]

    options = webdriver.ChromeOptions()
    options.add_argument("--start-maximized")
    driver = webdriver.Chrome(options=options)

    try:
        driver.get(WHATSAPP_WEB_URL)
        wait_for_login(driver)

        for row in range(2, sheet.max_row + 1):
            raw_phone = sheet.cell(row=row, column=1).value
            message = str(sheet.cell(row=row, column=2).value or "")
            attachment = str(sheet.cell(row=row, column=4).value or "").strip()

            phone = sanitize_phone_number(raw_phone, default_country_code)
            if not phone:
                status = "Skipped: invalid phone"
                print(f"Row {row}: {status}")
                write_result(sheet, row, status)
                workbook.save(excel_path)
                continue

            chat_text = message
            ok, reason = open_chat(driver, phone, chat_text)
            if not ok:
                status = f"Failed: {reason}"
                print(f"Row {row} ({phone}): {status}")
                write_result(sheet, row, status)
                workbook.save(excel_path)
                continue

            if attachment:
                media_ok, media_reason = send_attachment(driver, attachment, caption=message)
                final_status = "Sent attachment + caption" if media_ok else f"Failed: {media_reason}"
            else:
                txt_ok, txt_reason = send_text(driver)
                final_status = "Sent text" if txt_ok else f"Failed: {txt_reason}"

            print(f"Row {row} ({phone}): {final_status}")
            write_result(sheet, row, final_status)
            workbook.save(excel_path)

            time.sleep(max(0.0, pause_seconds))
            if cooldown_every > 0 and (row - 1) % cooldown_every == 0:
                print(f"Cooldown for {cooldown_seconds} seconds...")
                time.sleep(max(0.0, cooldown_seconds))

        print(f"Finished. Status written to {excel_path} columns E/F.")
    finally:
        workbook.save(excel_path)
        driver.quit()

    if open_excel_after:
        open_excel_file(excel_path)




def open_excel_file(path: Path) -> None:
    """Open output Excel file after run (best effort, mainly for Windows)."""
    try:
        if os.name == "nt":
            os.startfile(str(path))
    except Exception:
        pass

def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Send WhatsApp messages from Excel via WhatsApp Web.")
    parser.add_argument("excel", help="Path to Excel file (.xlsx or .xlsm)")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help=f"Sheet name (default: {DEFAULT_SHEET_NAME})")
    parser.add_argument("--default-country-code", default=None, help="Prefix for 10-digit local numbers, e.g. 91")
    parser.add_argument("--pause", type=float, default=3.0, help="Pause between rows in seconds")
    parser.add_argument("--cooldown-every", type=int, default=9, help="Run cooldown every N processed rows (0 to disable)")
    parser.add_argument("--cooldown-seconds", type=float, default=42.0, help="Cooldown duration in seconds")
    parser.add_argument("--open-excel-after", action="store_true", help="Open the Excel file after run (Windows)")
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    if not excel_path.exists():
        raise WhatsAppAutomationError(f"Excel file not found: {excel_path}")

    if excel_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise WhatsAppAutomationError("Use .xlsx or .xlsm file")

    process_rows(
        excel_path=excel_path,
        sheet_name=args.sheet,
        default_country_code=args.default_country_code,
        pause_seconds=args.pause,
        cooldown_every=args.cooldown_every,
        cooldown_seconds=args.cooldown_seconds,
        open_excel_after=args.open_excel_after,
    )


if __name__ == "__main__":
    main()
